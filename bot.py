from telebot.async_telebot import AsyncTeleBot
from telebot import types
from telebot.asyncio_handler_backends import State, StatesGroup
from telebot.asyncio_storage import StateMemoryStorage
import config
import json
import re
from datetime import datetime
from database import SessionLocal
from models.User import User
from models.Address import Address
from models.Project import Project
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import tempfile

logger = logging.getLogger(__name__)

# Initialize bot with state storage
state_storage = StateMemoryStorage()
bot = AsyncTeleBot(config.TOKEN, state_storage=state_storage)

# Define states for registration flow
class RegistrationStates(StatesGroup):
    full_name = State()
    region = State()
    district = State()
    mahalla = State()
    workplace = State()
    birth_date = State()
    passport_series = State()
    phone_number = State()
    confirmation = State()
    project_type = State()
    project_file = State()

welcome_message = """🌟 <b>"O'zbekiston — bag'rikeng diyor!" tanloviga xush kelibsiz!</b> 🇺🇿

Assalomu alaykum, aziz yoshlar! 

Din ishlari bo'yicha qo'mita tomonidan tashkil etilgan ushbu tanlovning maqsadi — yoshlar orasida <b>millatlar va dinlararo bag'rikenglikni mustahkamlash, radikal g'oyalarga qarshi kurashish, hamda tinchlik va totuvlik g'oyalarini targ'ib qilish</b>dir.

14 yoshdan 30 yoshgacha bo'lgan barcha ijodkor yoshlar o'z iste'dodini quyidagi yo'nalishlarda namoyish etishlari mumkin:
✍️ Maqola yoki esse
🎤 She'r yoki monolog
🎶 Qo'shiq yoki musiqiy chiqish
🎨 Rassomchilik ishi
🧵 Hunarmandchilik namunasi
🎥 Video-rolik yoki kontent

Bag'rikenglik, do'stlik va birlik g'oyalarini ifodalovchi ijodingizni bizga yuboring va tanlovda ishtirok eting!

Ro'yxatdan o'tish uchun tugmani bosing 👇"""

# Load regions data
def load_regions():
    try:
        with open("regions.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading regions: {e}")
        return {}

@bot.message_handler(commands=['start'])
async def welcome_handler(message: types.Message):
    """Handle /start command"""
    user_id = message.from_user.id
    username = message.from_user.username or "no_username"
    logger.info(f"User {user_id} (@{username}) started the bot with /start command")
    
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    
    # Check if user is admin
    if message.from_user.id in config.ADMIN_IDS:
        logger.info(f"Admin access granted for user {user_id}")
        markup.add(types.KeyboardButton("👤 Ro'yxatdan o'tish"))
        markup.add(types.KeyboardButton("📊 Ma'lumotlarni yuklab olish (Admin)"))
        admin_note = "\n\n🔐 <b>Admin panel mavjud!</b>"
    else:
        markup.add(types.KeyboardButton("👤 Ro'yxatdan o'tish"))
        admin_note = ""
    
    await bot.send_message(
        message.from_user.id,
        welcome_message + admin_note,
        parse_mode='HTML',
        reply_markup=markup
    )
    logger.info(f"Welcome message sent to user {user_id}")

@bot.message_handler(func=lambda message: message.text == "👤 Ro'yxatdan o'tish")
async def start_registration(message: types.Message):
    """Start registration process"""
    user_id = message.from_user.id
    logger.info(f"User {user_id} clicked registration button")
    
    # Check if user already exists
    db = SessionLocal()
    try:
        existing_user = db.query(User).filter(User.telegram_id == message.from_user.id).first()
        if existing_user:
            logger.info(f"User {user_id} already registered, showing options")
            # User already registered, show options
            markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
            markup.add(
                types.KeyboardButton("➕ Loyiha yuborish"),
                types.KeyboardButton("✏️ Ma'lumotlarni tahrirlash"),
                types.KeyboardButton("🏠 Bosh sahifa")
            )
            
            await bot.send_message(
                message.from_user.id,
                "✅ Siz allaqachon ro'yxatdan o'tgansiz!\n\n"
                "Quyidagi variantlardan birini tanlang:",
                reply_markup=markup
            )
            return
    finally:
        db.close()
    
    # Ask for full name
    logger.info(f"Starting new registration flow for user {user_id}")
    markup = types.ReplyKeyboardRemove()
    await bot.send_message(
        message.from_user.id,
        "📝 Iltimos, to'liq ismingizni kiriting:\n(Masalan: Aliyev Vali Akramovich)",
        reply_markup=markup
    )
    await bot.set_state(message.from_user.id, RegistrationStates.full_name, message.chat.id)

@bot.message_handler(state=RegistrationStates.full_name)
async def process_full_name(message: types.Message):
    """Process full name input"""
    # Check if user is editing
    is_editing = False
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        is_editing = 'phone_number' in data and data.get('phone_number') is not None
        data['full_name'] = message.text
    
    if is_editing:
        await show_confirmation(message)
        return
    
    # Show regions (normal flow)
    regions = load_regions()
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    
    for region_id, region_data in regions.items():
        markup.add(types.KeyboardButton(region_data['name']))
    
    await bot.send_message(
        message.from_user.id,
        "📍 Viloyatingizni tanlang:",
        reply_markup=markup
    )
    await bot.set_state(message.from_user.id, RegistrationStates.region, message.chat.id)

@bot.message_handler(state=RegistrationStates.region)
async def process_region(message: types.Message):
    """Process region selection"""
    regions = load_regions()
    selected_region = None
    
    # Find selected region
    for region_id, region_data in regions.items():
        if region_data['name'] == message.text:
            selected_region = region_id
            break
    
    if not selected_region:
        await bot.send_message(message.from_user.id, "❗️ Iltimos, tugmalardan birini tanlang!")
        return
    
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        data['region_id'] = selected_region
    
    # Show districts
    districts = regions[selected_region].get('districts', {})
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    
    for district_id, district_data in districts.items():
        markup.add(types.KeyboardButton(district_data['name']))
    
    await bot.send_message(
        message.from_user.id,
        "🏘 Tumaningizni tanlang:",
        reply_markup=markup
    )
    await bot.set_state(message.from_user.id, RegistrationStates.district, message.chat.id)

@bot.message_handler(state=RegistrationStates.district)
async def process_district(message: types.Message):
    """Process district selection"""
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        region_id = data.get('region_id')
    
    regions = load_regions()
    districts = regions[region_id].get('districts', {})
    selected_district = None
    
    # Find selected district
    for district_id, district_data in districts.items():
        if district_data['name'] == message.text:
            selected_district = district_id
            break
    
    if not selected_district:
        await bot.send_message(message.from_user.id, "❗️ Iltimos, tugmalardan birini tanlang!")
        return
    
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        data['district_id'] = selected_district
        data['district_name'] = message.text
    
    # Ask for mahalla
    markup = types.ReplyKeyboardRemove()
    await bot.send_message(
        message.from_user.id,
        "� Mahalla nomini kiriting:\n(Masalan: Yangi hayot mahallasi)",
        reply_markup=markup
    )
    await bot.set_state(message.from_user.id, RegistrationStates.mahalla, message.chat.id)

@bot.message_handler(state=RegistrationStates.mahalla)
async def process_mahalla(message: types.Message):
    """Process mahalla input"""
    # Check if user is editing
    is_editing = False
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        is_editing = 'phone_number' in data and data.get('phone_number') is not None
        data['mahalla'] = message.text
    
    if is_editing:
        await show_confirmation(message)
        return
    
    # Ask for workplace
    await bot.send_message(
        message.from_user.id,
        "🏢 Ish joyingizni kiriting:\n(Masalan: Toshkent davlat universiteti)"
    )
    await bot.set_state(message.from_user.id, RegistrationStates.workplace, message.chat.id)

@bot.message_handler(state=RegistrationStates.workplace)
async def process_workplace(message: types.Message):
    """Process workplace input"""
    # Check if user is editing
    is_editing = False
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        is_editing = 'phone_number' in data and data.get('phone_number') is not None
        data['workplace'] = message.text
    
    if is_editing:
        await show_confirmation(message)
        return
    
    await bot.send_message(
        message.from_user.id,
        "📅 Tug'ilgan sanangizni kiriting:\n(Masalan: 01.01.2000)"
    )
    await bot.set_state(message.from_user.id, RegistrationStates.birth_date, message.chat.id)

@bot.message_handler(state=RegistrationStates.birth_date)
async def process_birth_date(message: types.Message):
    """Process birth date input"""
    # Validate date format DD.MM.YYYY
    date_pattern = r'^\d{2}\.\d{2}\.\d{4}$'
    if not re.match(date_pattern, message.text):
        await bot.send_message(
            message.from_user.id,
            "❌ Noto'g'ri format! Iltimos, sanani DD.MM.YYYY formatida kiriting.\n(Masalan: 01.01.2000)"
        )
        return
    
    # Validate if it's a real date
    try:
        day, month, year = message.text.split('.')
        datetime(int(year), int(month), int(day))
    except ValueError:
        await bot.send_message(
            message.from_user.id,
            "❌ Noto'g'ri sana! Iltimos, mavjud sanani kiriting.\n(Masalan: 01.01.2000)"
        )
        return
    
    # Check if user is editing
    is_editing = False
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        is_editing = 'phone_number' in data and data.get('phone_number') is not None
        data['birth_date'] = message.text
    
    if is_editing:
        await show_confirmation(message)
        return
    
    await bot.send_message(
        message.from_user.id,
        "🆔 Pasport seriya va raqamingizni kiriting:\n(Masalan: AA1234567)"
    )
    await bot.set_state(message.from_user.id, RegistrationStates.passport_series, message.chat.id)

@bot.message_handler(state=RegistrationStates.passport_series)
async def process_passport(message: types.Message):
    """Process passport input"""
    # Validate passport format: 2 letters (any case) + 7 digits (e.g., AA1234567 or aa1234567)
    passport_pattern = r'^[A-Za-z]{2}\d{7}$'
    if not re.match(passport_pattern, message.text):
        await bot.send_message(
            message.from_user.id,
            "❌ Noto'g'ri format! Pasport seriyasi 2 ta harf va 7 ta raqamdan iborat bo'lishi kerak.\n(Masalan: AA1234567)"
        )
        return
    
    # Convert to uppercase before storing
    passport_upper = message.text.upper()
    
    # Check if user is editing
    is_editing = False
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        is_editing = 'phone_number' in data and data.get('phone_number') is not None
        data['passport_series'] = passport_upper
    
    if is_editing:
        await show_confirmation(message)
        return
    
    # Request phone number
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(types.KeyboardButton("📱 Telefon raqamni yuborish", request_contact=True))
    
    await bot.send_message(
        message.from_user.id,
        "📱 Telefon raqamingizni yuboring:",
        reply_markup=markup
    )
    await bot.set_state(message.from_user.id, RegistrationStates.phone_number, message.chat.id)

@bot.message_handler(state=RegistrationStates.phone_number, content_types=['contact', 'text'])
async def process_phone(message: types.Message):
    """Process phone number input"""
    phone_number = None
    
    if message.content_type == 'contact':
        phone_number = message.contact.phone_number
    else:
        phone_number = message.text
    
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        data['phone_number'] = phone_number
    
    # Show confirmation with all data
    await show_confirmation(message)

async def show_confirmation(message: types.Message):
    """Show user data for confirmation"""
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        # Load regions to get region and district names
        regions = load_regions()
        region_name = regions.get(data['region_id'], {}).get('name', 'N/A')
        district_name = data.get('district_name', 'N/A')
        
        confirmation_text = f"""
📋 <b>Ma'lumotlaringizni tekshiring:</b>

👤 <b>To'liq ism:</b> {data.get('full_name', 'N/A')}
📍 <b>Viloyat:</b> {region_name}
🏘 <b>Tuman:</b> {district_name}
🏘 <b>Mahalla:</b> {data.get('mahalla', 'N/A')}
🏢 <b>Ish joyi:</b> {data.get('workplace', 'N/A')}
📅 <b>Tug'ilgan sana:</b> {data.get('birth_date', 'N/A')}
🆔 <b>Pasport:</b> {data.get('passport_series', 'N/A')}
📱 <b>Telefon:</b> {data.get('phone_number', 'N/A')}

Ma'lumotlar to'g'rimi?
"""
    
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(
        types.KeyboardButton("✅ Ha, to'g'ri"),
        types.KeyboardButton("✏️ Tahrirlash")
    )
    
    await bot.send_message(
        message.from_user.id,
        confirmation_text,
        parse_mode='HTML',
        reply_markup=markup
    )
    await bot.set_state(message.from_user.id, RegistrationStates.confirmation, message.chat.id)

@bot.message_handler(state=RegistrationStates.confirmation)
async def process_confirmation(message: types.Message):
    """Process confirmation response"""
    user_id = message.from_user.id
    logger.info(f"Confirmation handler received: '{message.text}' from user {user_id}")
    
    if message.text == "✅ Ha, to'g'ri":
        # Check if user already exists (editing existing data)
        db = SessionLocal()
        try:
            existing_user = db.query(User).filter(User.telegram_id == message.from_user.id).first()
            
            if existing_user:
                logger.info(f"Updating existing user data for user {user_id}")
                # Update existing user data
                async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
                    # Update address
                    address = db.query(Address).filter(Address.id == existing_user.address_id).first()
                    if address:
                        address.region_id = int(data['region_id'])
                        address.district_id = int(data['district_id'])
                        address.neighborhood = data.get('mahalla', '')
                    
                    # Update user
                    existing_user.full_name = data['full_name']
                    existing_user.workplace = data['workplace']
                    existing_user.birth_date = data['birth_date']
                    existing_user.passport_series = data['passport_series']
                    existing_user.phone_number = data['phone_number']
                    
                    db.commit()
                    logger.info(f"User {user_id} data updated successfully")
                    
                    # Show success and options
                    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
                    markup.add(
                        types.KeyboardButton("➕ Loyiha yuborish"),
                        types.KeyboardButton("🏠 Bosh sahifa")
                    )
                    
                    await bot.send_message(
                        message.from_user.id,
                        "✅ Ma'lumotlaringiz muvaffaqiyatli yangilandi!",
                        reply_markup=markup
                    )
                    
                    # Clear state
                    await bot.delete_state(message.from_user.id, message.chat.id)
                    return
                    
        finally:
            db.close()
        
        # If user doesn't exist, continue to project submission
        # Ask for project type
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        for project_key, project_info in config.PROJECT_TYPES.items():
            markup.add(types.KeyboardButton(project_info['title']))
        
        await bot.send_message(
            message.from_user.id,
            "🎨 Loyiha turini tanlang:",
            reply_markup=markup
        )
        await bot.set_state(message.from_user.id, RegistrationStates.project_type, message.chat.id)
    
    elif message.text == "✏️ Tahrirlash":
        # Show edit options
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        markup.add(
            types.KeyboardButton("👤 Ism"),
            types.KeyboardButton("📍 Manzil"),
            types.KeyboardButton("🏢 Ish joyi"),
            types.KeyboardButton("📅 Tug'ilgan sana"),
            types.KeyboardButton("🆔 Pasport"),
            types.KeyboardButton("📱 Telefon"),
            types.KeyboardButton("🔙 Orqaga")
        )
        
        await bot.send_message(
            message.from_user.id,
            "✏️ Qaysi ma'lumotni o'zgartirmoqchisiz?",
            reply_markup=markup
        )
        # Stay in confirmation state to handle edit buttons
    
    elif message.text == "🔙 Orqaga":
        await show_confirmation(message)
    
    elif message.text == "👤 Ism":
        markup = types.ReplyKeyboardRemove()
        await bot.send_message(
            message.from_user.id,
            "📝 Yangi ismingizni kiriting:",
            reply_markup=markup
        )
        await bot.set_state(message.from_user.id, RegistrationStates.full_name, message.chat.id)
    
    elif message.text == "📍 Manzil":
        regions = load_regions()
        region_markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        for region_id, region_data in regions.items():
            region_markup.add(types.KeyboardButton(region_data['name']))
        
        await bot.send_message(
            message.from_user.id,
            "📍 Viloyatingizni tanlang:",
            reply_markup=region_markup
        )
        await bot.set_state(message.from_user.id, RegistrationStates.region, message.chat.id)
    
    elif message.text == "🏢 Ish joyi":
        markup = types.ReplyKeyboardRemove()
        await bot.send_message(
            message.from_user.id,
            "🏢 Yangi ish joyingizni kiriting:",
            reply_markup=markup
        )
        await bot.set_state(message.from_user.id, RegistrationStates.workplace, message.chat.id)
    
    elif message.text == "📅 Tug'ilgan sana":
        markup = types.ReplyKeyboardRemove()
        await bot.send_message(
            message.from_user.id,
            "📅 Yangi tug'ilgan sanangizni kiriting (DD.MM.YYYY):",
            reply_markup=markup
        )
        await bot.set_state(message.from_user.id, RegistrationStates.birth_date, message.chat.id)
    
    elif message.text == "🆔 Pasport":
        markup = types.ReplyKeyboardRemove()
        await bot.send_message(
            message.from_user.id,
            "🆔 Yangi pasport ma'lumotingizni kiriting (AA1234567):",
            reply_markup=markup
        )
        await bot.set_state(message.from_user.id, RegistrationStates.passport_series, message.chat.id)
    
    elif message.text == "📱 Telefon":
        phone_markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        phone_markup.add(types.KeyboardButton("📱 Telefon raqamni yuborish", request_contact=True))
        
        await bot.send_message(
            message.from_user.id,
            "📱 Yangi telefon raqamingizni yuboring:",
            reply_markup=phone_markup
        )
        await bot.set_state(message.from_user.id, RegistrationStates.phone_number, message.chat.id)
    
    else:
        await bot.send_message(
            message.from_user.id,
            "❗️ Iltimos, tugmalardan birini tanlang!"
        )

@bot.message_handler(state=RegistrationStates.project_type)
async def process_project_type(message: types.Message):
    """Process project type selection"""
    user_id = message.from_user.id
    logger.info(f"Processing project type selection from user {user_id}: '{message.text}'")
    
    # Find selected project type
    selected_type = None
    for project_key, project_info in config.PROJECT_TYPES.items():
        if project_info['title'] == message.text:
            selected_type = project_key
            break
    
    if not selected_type:
        logger.warning(f"User {user_id} selected invalid project type: '{message.text}'")
        await bot.send_message(message.from_user.id, "❗️ Iltimos, tugmalardan birini tanlang!")
        return
    
    logger.info(f"User {user_id} selected project type: {selected_type}")
    
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        data['project_type'] = selected_type
    
    # Ask for project file
    markup = types.ReplyKeyboardRemove()
    file_types = config.PROJECT_TYPES[selected_type]['file_types']
    await bot.send_message(
        message.from_user.id,
        f"📎 Loyihangizni yuklang:\n\n"
        f"Qo'llab-quvvatlanadigan formatlar: {file_types}\n\n"
        f"Faylni shu yerga yuboring.",
        reply_markup=markup
    )
    
    # Set state to project_file
    await bot.set_state(message.from_user.id, RegistrationStates.project_file, message.chat.id)
    
    # Verify state was set
    current_state = await bot.get_state(message.from_user.id, message.chat.id)
    logger.info(f"State set for user {user_id}: '{current_state}' (expected: '{RegistrationStates.project_file.name}')")

@bot.message_handler(state=RegistrationStates.project_file, content_types=['document', 'photo', 'audio', 'video', 'voice'])
async def process_project_file(message: types.Message):
    """Process project file submission"""
    user_id = message.from_user.id
    content_type = message.content_type
    logger.info(f"Received project file ({content_type}) from user {user_id}")
    
    async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
        project_type = data.get('project_type')
        
        # Save to database first
        db = SessionLocal()
        try:
            # Check if user exists, if not create new user
            user = db.query(User).filter(User.telegram_id == message.from_user.id).first()
            
            if not user:
                # Create address
                address = Address(
                    region_id=int(data['region_id']),
                    district_id=int(data['district_id']),
                    neighborhood=data.get('mahalla', '')
                )
                db.add(address)
                db.flush()
                
                # Create user
                user = User(
                    telegram_id=message.from_user.id,
                    full_name=data['full_name'],
                    address_id=address.id,
                    workplace=data['workplace'],
                    birth_date=data['birth_date'],
                    passport_series=data['passport_series'],
                    phone_number=data['phone_number']
                )
                db.add(user)
                db.flush()
            
            # Load user's address from database
            address = db.query(Address).filter(Address.id == user.address_id).first()
            
            # Forward the project file to the channel
            try:
                # Forward the file to channel
                forwarded_msg = await bot.forward_message(
                    chat_id=config.CHANNEL_ID,
                    from_chat_id=message.chat.id,
                    message_id=message.message_id
                )
                
                # Prepare user data message from database
                regions = load_regions()
                region_name = "N/A"
                district_name = "N/A"
                mahalla = "N/A"
                
                if address:
                    region_name = regions.get(str(address.region_id), {}).get('name', 'N/A')
                    districts = regions.get(str(address.region_id), {}).get('districts', {})
                    district_name = districts.get(str(address.district_id), {}).get('name', 'N/A')
                    mahalla = address.neighborhood or "N/A"
                
                project_type_title = config.PROJECT_TYPES.get(project_type, {}).get('title', 'N/A')
                
                user_data_text = f"""
📋 <b>Ishtirokchi ma'lumotlari:</b>

👤 <b>Ism:</b> {user.full_name or 'N/A'}
📍 <b>Viloyat:</b> {region_name}
🏘 <b>Tuman:</b> {district_name}
🏘 <b>Mahalla:</b> {mahalla}
🏢 <b>Ish joyi:</b> {user.workplace or 'N/A'}
📅 <b>Tug'ilgan sana:</b> {user.birth_date or 'N/A'}
🆔 <b>Pasport:</b> {user.passport_series or 'N/A'}
📱 <b>Telefon:</b> {user.phone_number or 'N/A'}
🎨 <b>Loyiha turi:</b> {project_type_title}
"""
                
                # Send user data as reply to the forwarded message
                user_data_msg = await bot.send_message(
                    chat_id=config.CHANNEL_ID,
                    text=user_data_text,
                    parse_mode='HTML',
                    reply_to_message_id=forwarded_msg.message_id
                )
                
                # Get the message URL (construct it from channel and message id)
                # Format for private channels: https://t.me/c/{channel_id_without_-100}/{message_id}
                # Convert channel ID: -1003119110887 -> 3119110887
                channel_id_str = str(config.CHANNEL_ID)
                if channel_id_str.startswith('-100'):
                    channel_id_clean = channel_id_str[4:]  # Remove '-100' prefix
                    project_url = f"https://t.me/c/{channel_id_clean}/{forwarded_msg.message_id}"
                else:
                    # For public channels with @ username
                    channel_username = config.CHANNEL_ID.replace('@', '')
                    project_url = f"https://t.me/{channel_username}/{forwarded_msg.message_id}"
                
                # Create project record
                project = Project(
                    user_id=user.id,
                    type=project_type,
                    project_url=project_url
                )
                db.add(project)
                db.commit()
                
                logger.info(f"Project saved successfully for user {user_id}: type={project_type}, url={project_url}")
                
                # Success message with option to submit another project
                markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
                markup.add(
                    types.KeyboardButton("➕ Yana loyiha yuborish"),
                    types.KeyboardButton("🏠 Bosh sahifa")
                )
                
                await bot.send_message(
                    message.from_user.id,
                    "✅ Rahmat! Loyihangiz muvaffaqiyatli yuborildi.\n\n"
                    "Sizning loyihangiz ko'rib chiqiladi va natijalar keyinroq e'lon qilinadi.",
                    reply_markup=markup
                )
                
            except Exception as e:
                logger.error(f"Error forwarding to channel: {e}")
                db.rollback()
                await bot.send_message(
                    message.from_user.id,
                    "❌ Loyihani yuborishda xatolik yuz berdi. Iltimos, qaytadan urinib ko'ring."
                )
                return
                
        except Exception as e:
            logger.error(f"Error saving user data: {e}", exc_info=True)
            db.rollback()
            await bot.send_message(
                message.from_user.id,
                "❌ Xatolik yuz berdi. Iltimos, qaytadan urinib ko'ring."
            )
        finally:
            db.close()
    
    # Delete state
    await bot.delete_state(message.from_user.id, message.chat.id)

@bot.message_handler(func=lambda message: message.text == "🏠 Bosh sahifa")
async def go_home(message: types.Message):
    """Return to home page"""
    await welcome_handler(message)

@bot.message_handler(func=lambda message: message.text in ["➕ Yana loyiha yuborish", "➕ Loyiha yuborish"])
async def submit_another_project(message: types.Message):
    """Allow user to submit another project"""
    user_id = message.from_user.id
    logger.info(f"User {user_id} clicked submit project button: '{message.text}'")
    
    # Ask for project type
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    for project_key, project_info in config.PROJECT_TYPES.items():
        markup.add(types.KeyboardButton(project_info['title']))
    
    await bot.send_message(
        message.from_user.id,
        "🎨 Loyiha turini tanlang:",
        reply_markup=markup
    )
    await bot.set_state(message.from_user.id, RegistrationStates.project_type, message.chat.id)
    logger.info(f"State set to project_type for user {user_id}")

@bot.message_handler(func=lambda message: message.text == "✏️ Ma'lumotlarni tahrirlash")
async def edit_personal_info(message: types.Message):
    """Allow user to edit their personal information"""
    # Check if user exists
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.telegram_id == message.from_user.id).first()
        if not user:
            await bot.send_message(
                message.from_user.id,
                "❌ Siz hali ro'yxatdan o'tmagansiz. Iltimos, avval ro'yxatdan o'ting."
            )
            return
        
        # Get user's address
        address = None
        if user.address_id:
            address = db.query(Address).filter(Address.id == user.address_id).first()
        
        # Set state first to initialize storage
        await bot.set_state(message.from_user.id, RegistrationStates.confirmation, message.chat.id)
        
        # Load user data into state
        async with bot.retrieve_data(message.from_user.id, message.chat.id) as data:
            data['full_name'] = user.full_name or ''
            data['region_id'] = str(address.region_id) if address and address.region_id else '1'
            data['district_id'] = str(address.district_id) if address and address.district_id else '1'
            data['mahalla'] = address.neighborhood if address and address.neighborhood else ''
            data['workplace'] = user.workplace or ''
            data['birth_date'] = user.birth_date or ''
            data['passport_series'] = user.passport_series or ''
            data['phone_number'] = user.phone_number or ''
            
            # Get district name for confirmation display
            if address and address.region_id and address.district_id:
                try:
                    regions = load_regions()
                    districts = regions.get(str(address.region_id), {}).get('districts', {})
                    data['district_name'] = districts.get(str(address.district_id), {}).get('name', 'N/A')
                except Exception as e:
                    logger.error(f"Error loading district name: {e}")
                    data['district_name'] = 'N/A'
            else:
                data['district_name'] = 'N/A'
        
        # Show confirmation screen
        await show_confirmation(message)
        
    except Exception as e:
        logger.error(f"Error in edit_personal_info: {e}", exc_info=True)
        await bot.send_message(
            message.from_user.id,
            "❌ Xatolik yuz berdi. Iltimos, qaytadan urinib ko'ring."
        )
    finally:
        db.close()

# Handle contact sharing separately
@bot.message_handler(content_types=['contact'])
async def handle_contact(message: types.Message):
    """Handle contact sharing"""
    current_state = await bot.get_state(message.from_user.id, message.chat.id)
    logger.info(f"Contact received from user {message.from_user.id} in state {current_state}")
    
    if current_state == RegistrationStates.phone_number.name:
        await process_phone(message)
    else:
        await bot.send_message(
            message.from_user.id,
            "❌ Kutilmagan harakat. Iltimos, /start dan boshlang."
        )

# Handle file uploads separately
@bot.message_handler(content_types=['document', 'photo', 'audio', 'video', 'voice'])
async def handle_file_upload(message: types.Message):
    """Handle file uploads"""
    current_state = await bot.get_state(message.from_user.id, message.chat.id)
    logger.info(f"File received from user {message.from_user.id} in state '{current_state}', content_type: {message.content_type}")
    logger.info(f"Expected state: '{RegistrationStates.project_file.name}'")
    logger.info(f"State match: {current_state == RegistrationStates.project_file.name}")
    
    if current_state == RegistrationStates.project_file.name:
        logger.info(f"Routing to process_project_file for user {message.from_user.id}")
        await process_project_file(message)
    else:
        logger.warning(f"User {message.from_user.id} sent file without being in project_file state. Current state: {current_state}")
        await bot.send_message(
            message.from_user.id,
            f"❌ Iltimos, avval loyiha turini tanlang.\n\n"
            f"Qadam:\n"
            f"1️⃣ \"➕ Loyiha yuborish\" tugmasini bosing\n"
            f"2️⃣ Loyiha turini tanlang\n"
            f"3️⃣ Faylni yuboring"
        )

@bot.message_handler(func=lambda message: message.text == "📊 Ma'lumotlarni yuklab olish (Admin)")
async def export_data_admin(message: types.Message):
    """Export all data to Excel (Admin only)"""
    user_id = message.from_user.id
    # Check if user is admin
    if user_id not in config.ADMIN_IDS:
        logger.warning(f"Non-admin user {user_id} attempted to access admin export function")
        await bot.send_message(
            message.from_user.id,
            "❌ Bu funksiya faqat adminlar uchun!"
        )
        return
    
    logger.info(f"Admin {user_id} initiated data export")
    await bot.send_message(
        message.from_user.id,
        "⏳ Ma'lumotlar tayyorlanmoqda, iltimos kuting..."
    )
    
    db = SessionLocal()
    try:
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: All Users with their data
        ws_users = wb.active
        ws_users.title = "Foydalanuvchilar"
        
        # Header styling - Enhanced colors
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Data styling
        even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Light gray
        data_font = Font(size=11)
        
        # Border styling
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Headers for users sheet
        headers = [
            "№", "Telegram ID", "To'liq ism", "Viloyat", "Tuman", "Mahalla",
            "Ish joyi", "Tug'ilgan sana", "Pasport", "Telefon", "Loyihalar soni"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_users.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Load regions for display
        regions = load_regions()
        
        # Fetch all users with their addresses
        users = db.query(User).all()
        
        for idx, user in enumerate(users, 2):
            address = db.query(Address).filter(Address.id == user.address_id).first()
            
            # Get region and district names
            region_name = "N/A"
            district_name = "N/A"
            mahalla = "N/A"
            
            if address:
                region_name = regions.get(str(address.region_id), {}).get('name', 'N/A')
                districts = regions.get(str(address.region_id), {}).get('districts', {})
                district_name = districts.get(str(address.district_id), {}).get('name', 'N/A')
                mahalla = address.neighborhood or "N/A"
            
            # Count projects
            project_count = db.query(Project).filter(Project.user_id == user.id).count()
            
            # Write user data
            ws_users.cell(row=idx, column=1).value = idx - 1
            ws_users.cell(row=idx, column=2).value = user.telegram_id
            ws_users.cell(row=idx, column=3).value = user.full_name or "N/A"
            ws_users.cell(row=idx, column=4).value = region_name
            ws_users.cell(row=idx, column=5).value = district_name
            ws_users.cell(row=idx, column=6).value = mahalla
            ws_users.cell(row=idx, column=7).value = user.workplace or "N/A"
            ws_users.cell(row=idx, column=8).value = user.birth_date or "N/A"
            ws_users.cell(row=idx, column=9).value = user.passport_series or "N/A"
            ws_users.cell(row=idx, column=10).value = user.phone_number or "N/A"
            ws_users.cell(row=idx, column=11).value = project_count
            
            # Apply styling to all cells in row
            for col in range(1, 12):
                cell = ws_users.cell(row=idx, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = data_font
                cell.border = thin_border
                # Alternate row colors for better readability
                if idx % 2 == 0:
                    cell.fill = even_row_fill
        
        # Set specific column widths for better layout
        ws_users.column_dimensions['A'].width = 6   # №
        ws_users.column_dimensions['B'].width = 14  # Telegram ID
        ws_users.column_dimensions['C'].width = 30  # Full name
        ws_users.column_dimensions['D'].width = 20  # Region
        ws_users.column_dimensions['E'].width = 20  # District
        ws_users.column_dimensions['F'].width = 25  # Mahalla
        ws_users.column_dimensions['G'].width = 30  # Workplace
        ws_users.column_dimensions['H'].width = 15  # Birth date
        ws_users.column_dimensions['I'].width = 15  # Passport
        ws_users.column_dimensions['J'].width = 16  # Phone
        ws_users.column_dimensions['K'].width = 12  # Projects count
        
        # Freeze the header row
        ws_users.freeze_panes = 'A2'
        
        # Sheet 2: All Projects
        ws_projects = wb.create_sheet("Loyihalar")
        
        # Headers for projects sheet
        project_headers = [
            "№", "Ishtirokchi", "Telegram ID", "Loyiha turi", "Loyiha URL",
            "Viloyat", "Tuman", "Telefon"
        ]
        
        for col_num, header in enumerate(project_headers, 1):
            cell = ws_projects.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        
        # Fetch all projects
        projects = db.query(Project).all()
        
        for idx, project in enumerate(projects, 2):
            user = db.query(User).filter(User.id == project.user_id).first()
            address = db.query(Address).filter(Address.id == user.address_id).first() if user else None
            
            # Get region and district names
            region_name = "N/A"
            district_name = "N/A"
            
            if address:
                region_name = regions.get(str(address.region_id), {}).get('name', 'N/A')
                districts = regions.get(str(address.region_id), {}).get('districts', {})
                district_name = districts.get(str(address.district_id), {}).get('name', 'N/A')
            
            # Get project type title
            project_type_title = config.PROJECT_TYPES.get(project.type, {}).get('title', project.type)
            
            # Write project data
            ws_projects.cell(row=idx, column=1).value = idx - 1
            ws_projects.cell(row=idx, column=2).value = user.full_name if user else "N/A"
            ws_projects.cell(row=idx, column=3).value = user.telegram_id if user else "N/A"
            ws_projects.cell(row=idx, column=4).value = project_type_title
            
            # Make URL clickable with hyperlink
            url_cell = ws_projects.cell(row=idx, column=5)
            if project.project_url:
                url_cell.value = project.project_url
                url_cell.hyperlink = project.project_url
                url_cell.font = Font(color="0563C1", underline="single", size=11)  # Blue and underlined
            else:
                url_cell.value = "N/A"
            
            ws_projects.cell(row=idx, column=6).value = region_name
            ws_projects.cell(row=idx, column=7).value = district_name
            ws_projects.cell(row=idx, column=8).value = user.phone_number if user else "N/A"
            
            # Apply styling to all cells in row
            for col in range(1, 9):
                cell = ws_projects.cell(row=idx, column=col)
                cell.border = thin_border
                if col != 5:  # Don't center the URL
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = data_font
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Alternate row colors
                if idx % 2 == 0 and col != 5:  # Don't color the URL cell
                    cell.fill = even_row_fill
        
        # Set specific column widths for projects sheet
        ws_projects.column_dimensions['A'].width = 6   # №
        ws_projects.column_dimensions['B'].width = 30  # Participant name
        ws_projects.column_dimensions['C'].width = 14  # Telegram ID
        ws_projects.column_dimensions['D'].width = 25  # Project type
        ws_projects.column_dimensions['E'].width = 45  # Project URL (wider for links)
        ws_projects.column_dimensions['F'].width = 20  # Region
        ws_projects.column_dimensions['G'].width = 20  # District
        ws_projects.column_dimensions['H'].width = 16  # Phone
        
        # Freeze the header row
        ws_projects.freeze_panes = 'A2'
        
        # Sheet 3: Statistics
        ws_stats = wb.create_sheet("Statistika")
        
        total_users = len(users)
        total_projects = len(projects)
        
        # Count by regions
        region_stats = {}
        for user in users:
            address = db.query(Address).filter(Address.id == user.address_id).first()
            if address:
                region_name = regions.get(str(address.region_id), {}).get('name', 'Noma\'lum')
                region_stats[region_name] = region_stats.get(region_name, 0) + 1
        
        # Count by project types
        project_type_stats = {}
        for project in projects:
            project_type_title = config.PROJECT_TYPES.get(project.type, {}).get('title', project.type)
            project_type_stats[project_type_title] = project_type_stats.get(project_type_title, 0) + 1
        
        # Write statistics with enhanced styling
        row = 1
        
        # Main title
        title_cell = ws_stats.cell(row=row, column=1)
        title_cell.value = "UMUMIY STATISTIKA"
        title_cell.font = Font(bold=True, size=16, color="1F4E78")
        title_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws_stats.merge_cells(f'A{row}:B{row}')
        ws_stats.row_dimensions[row].height = 25
        row += 2
        
        # Summary statistics
        summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        label_cell = ws_stats.cell(row=row, column=1)
        label_cell.value = "Jami ro'yxatdan o'tganlar:"
        label_cell.font = Font(bold=True, size=12)
        label_cell.fill = summary_fill
        value_cell = ws_stats.cell(row=row, column=2)
        value_cell.value = total_users
        value_cell.font = Font(bold=True, size=12, color="C00000")
        value_cell.fill = summary_fill
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1
        
        label_cell = ws_stats.cell(row=row, column=1)
        label_cell.value = "Jami yuborilgan loyihalar:"
        label_cell.font = Font(bold=True, size=12)
        label_cell.fill = summary_fill
        value_cell = ws_stats.cell(row=row, column=2)
        value_cell.value = total_projects
        value_cell.font = Font(bold=True, size=12, color="C00000")
        value_cell.fill = summary_fill
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 3
        
        # Region statistics
        section_cell = ws_stats.cell(row=row, column=1)
        section_cell.value = "VILOYATLAR BO'YICHA"
        section_cell.font = Font(bold=True, size=13, color="FFFFFF")
        section_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_stats.merge_cells(f'A{row}:B{row}')
        ws_stats.row_dimensions[row].height = 20
        row += 1
        
        # Region data with alternating colors
        for idx_stat, (region_name, count) in enumerate(sorted(region_stats.items(), key=lambda x: x[1], reverse=True), 1):
            name_cell = ws_stats.cell(row=row, column=1)
            name_cell.value = region_name
            name_cell.font = Font(size=11)
            count_cell = ws_stats.cell(row=row, column=2)
            count_cell.value = count
            count_cell.font = Font(size=11, bold=True)
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if idx_stat % 2 == 0:
                name_cell.fill = even_row_fill
                count_cell.fill = even_row_fill
            row += 1
        
        row += 2
        
        # Project type statistics
        section_cell = ws_stats.cell(row=row, column=1)
        section_cell.value = "LOYIHA TURLARI BO'YICHA"
        section_cell.font = Font(bold=True, size=13, color="FFFFFF")
        section_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_stats.merge_cells(f'A{row}:B{row}')
        ws_stats.row_dimensions[row].height = 20
        row += 1
        
        # Project type data with alternating colors
        for idx_stat, (project_type, count) in enumerate(sorted(project_type_stats.items(), key=lambda x: x[1], reverse=True), 1):
            name_cell = ws_stats.cell(row=row, column=1)
            name_cell.value = project_type
            name_cell.font = Font(size=11)
            count_cell = ws_stats.cell(row=row, column=2)
            count_cell.value = count
            count_cell.font = Font(size=11, bold=True)
            count_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            if idx_stat % 2 == 0:
                name_cell.fill = even_row_fill
                count_cell.fill = even_row_fill
            row += 1
        
        # Set column widths for stats sheet
        ws_stats.column_dimensions['A'].width = 45
        ws_stats.column_dimensions['B'].width = 18
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Send file to admin
        with open(temp_file.name, 'rb') as file:
            await bot.send_document(
                message.from_user.id,
                file,
                caption=f"📊 <b>Tanlov ma'lumotlari</b>\n\n"
                        f"👥 Jami ishtirokchilar: {total_users}\n"
                        f"📁 Jami loyihalar: {total_projects}\n"
                        f"📅 Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                parse_mode='HTML',
                visible_file_name=f"Tanlov_malumotlari_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            )
        
        # Delete temporary file
        os.unlink(temp_file.name)
        
        logger.info(f"Admin {user_id} exported data successfully: {total_users} users, {total_projects} projects")
        
    except Exception as e:
        logger.error(f"Error exporting data: {e}", exc_info=True)
        await bot.send_message(
            message.from_user.id,
            "❌ Ma'lumotlarni yuklashda xatolik yuz berdi. Iltimos, qaytadan urinib ko'ring."
        )
    finally:
        db.close()

# Debug: Catch-all handler to see unhandled messages
@bot.message_handler(func=lambda message: True, content_types=['text'])
async def debug_handler(message: types.Message):
    """Catch-all handler to route unhandled messages"""
    current_state = await bot.get_state(message.from_user.id, message.chat.id)
    
    # Route to correct handler based on state
    if current_state == RegistrationStates.full_name.name:
        await process_full_name(message)
    elif current_state == RegistrationStates.region.name:
        await process_region(message)
    elif current_state == RegistrationStates.district.name:
        await process_district(message)
    elif current_state == RegistrationStates.mahalla.name:
        await process_mahalla(message)
    elif current_state == RegistrationStates.workplace.name:
        await process_workplace(message)
    elif current_state == RegistrationStates.birth_date.name:
        await process_birth_date(message)
    elif current_state == RegistrationStates.passport_series.name:
        await process_passport(message)
    elif current_state == RegistrationStates.phone_number.name:
        await process_phone(message)
    elif current_state == RegistrationStates.confirmation.name:
        await process_confirmation(message)
    elif current_state == RegistrationStates.project_type.name:
        await process_project_type(message)
